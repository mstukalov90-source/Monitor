"""Poll excel_inbox and fill odh_export.ogh_analiz.ozn_date / executor."""

from __future__ import annotations

import hashlib
import logging
import time
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Sequence

from openpyxl import load_workbook
from psycopg2.extras import execute_values

from collector.config import EXCEL_UPLOAD_DIR, OGH_ANALIZ_SQL, OZN_UPLOAD_LOG_SQL
from collector.db import execute_sql_file, local_connection, log_job_run

logger = logging.getLogger(__name__)

JOB_NAME = "ozn_excel_inbox"
EXCEL_SUFFIXES = {".xlsx", ".xls"}
TEMP_NAME_PREFIXES = ("~$",)
TEMP_SUFFIXES = (".tmp", ".part", ".crdownload")
MIN_AGE_SECONDS = 2.0


@dataclass(frozen=True)
class ExcelRow:
    order_no: str
    ozn_date: date
    executor: str


@dataclass
class ParseResult:
    rows: list[ExcelRow]
    skipped_rows: int = 0


@dataclass
class FileProcessResult:
    file_name: str
    status: str
    excel_rows: int = 0
    filled_ordername: int = 0
    filled_order: int = 0
    missing_count: int = 0
    skipped_rows: int = 0
    missing_orders: list[str] = field(default_factory=list)
    error_message: str | None = None
    duration_ms: int = 0


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _as_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel

            parsed = from_excel(value)
        except Exception:
            return None
        if isinstance(parsed, datetime):
            return parsed.date()
        if isinstance(parsed, date):
            return parsed
        return None
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def parse_excel(path: Path) -> ParseResult:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        next(rows_iter, None)
        by_key: dict[str, ExcelRow] = {}
        skipped = 0
        for raw in rows_iter:
            if raw is None or len(raw) < 3:
                skipped += 1
                continue
            order_raw, date_raw, executor_raw = raw[0], raw[1], raw[2]
            if order_raw is None or str(order_raw).strip() == "":
                skipped += 1
                continue
            order_no = str(order_raw).strip()
            ozn_date = _as_date(date_raw)
            executor = "" if executor_raw is None else str(executor_raw).strip()
            if ozn_date is None or not executor:
                skipped += 1
                continue
            by_key[order_no] = ExcelRow(order_no, ozn_date, executor)
        return ParseResult(rows=list(by_key.values()), skipped_rows=skipped)
    finally:
        workbook.close()


def _is_temp_name(path: Path) -> bool:
    name = path.name
    if name.startswith(TEMP_NAME_PREFIXES):
        return True
    lowered = name.lower()
    return any(lowered.endswith(suffix) for suffix in TEMP_SUFFIXES)


def list_ready_files(inbox: Path, *, now: float | None = None) -> list[Path]:
    if not inbox.is_dir():
        return []
    clock = time.time() if now is None else now
    ready: list[Path] = []
    for path in sorted(inbox.iterdir()):
        if not path.is_file():
            continue
        if path.suffix.lower() not in EXCEL_SUFFIXES:
            continue
        if _is_temp_name(path):
            continue
        try:
            age = clock - path.stat().st_mtime
        except OSError:
            continue
        if age < MIN_AGE_SECONDS:
            continue
        ready.append(path)
    return ready


def _ensure_tables(conn) -> None:
    if not OGH_ANALIZ_SQL.exists():
        raise FileNotFoundError(f"SQL migration not found: {OGH_ANALIZ_SQL}")
    if not OZN_UPLOAD_LOG_SQL.exists():
        raise FileNotFoundError(f"SQL migration not found: {OZN_UPLOAD_LOG_SQL}")
    execute_sql_file(conn, OGH_ANALIZ_SQL)
    execute_sql_file(conn, OZN_UPLOAD_LOG_SQL)


def _apply_rows(cur, rows: Sequence[ExcelRow]) -> tuple[int, int, list[str]]:
    if not rows:
        return 0, 0, []

    cur.execute(
        """
        CREATE TEMP TABLE ozn_src (
            order_no text PRIMARY KEY,
            ozn_date date NOT NULL,
            executor text NOT NULL
        ) ON COMMIT DROP
        """
    )
    execute_values(
        cur,
        "INSERT INTO ozn_src (order_no, ozn_date, executor) VALUES %s",
        [(row.order_no, row.ozn_date, row.executor) for row in rows],
    )

    cur.execute(
        """
        UPDATE odh_export.ogh_analiz AS t
        SET ozn_date = s.ozn_date,
            executor = s.executor
        FROM ozn_src AS s
        WHERE t."OrderName" = s.order_no
        RETURNING s.order_no
        """
    )
    matched_ordername = {item[0] for item in cur.fetchall()}

    cur.execute(
        """
        UPDATE odh_export.ogh_analiz AS t
        SET ozn_date = s.ozn_date,
            executor = s.executor
        FROM ozn_src AS s
        WHERE t."order" = s.order_no
          AND NOT EXISTS (
              SELECT 1 FROM odh_export.ogh_analiz x
              WHERE x."OrderName" = s.order_no
          )
        RETURNING s.order_no
        """
    )
    matched_order = {item[0] for item in cur.fetchall()}

    missing = [
        row.order_no
        for row in rows
        if row.order_no not in matched_ordername and row.order_no not in matched_order
    ]
    return len(matched_ordername), len(matched_order), missing


def _insert_log(
    cur,
    *,
    file_name: str,
    file_hash: str | None,
    file_size: int | None,
    result: FileProcessResult,
) -> None:
    cur.execute(
        """
        INSERT INTO odh_export.ozn_upload_log (
            file_name,
            file_sha256,
            file_size_bytes,
            status,
            excel_rows,
            filled_ordername,
            filled_order,
            missing_count,
            skipped_rows,
            missing_orders,
            error_message,
            duration_ms
        ) VALUES (
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
        )
        """,
        (
            file_name,
            file_hash,
            file_size,
            result.status,
            result.excel_rows,
            result.filled_ordername,
            result.filled_order,
            result.missing_count,
            result.skipped_rows,
            result.missing_orders or None,
            result.error_message,
            result.duration_ms,
        ),
    )


def process_file(path: Path) -> FileProcessResult:
    started = time.monotonic()
    file_name = path.name
    file_hash: str | None = None
    file_size: int | None = None
    result = FileProcessResult(file_name=file_name, status="failed")

    try:
        file_size = path.stat().st_size
        file_hash = file_sha256(path)
        parsed = parse_excel(path)
        with local_connection() as conn:
            _ensure_tables(conn)
            with conn.cursor() as cur:
                filled_ordername, filled_order, missing = _apply_rows(cur, parsed.rows)
                result = FileProcessResult(
                    file_name=file_name,
                    status="success",
                    excel_rows=len(parsed.rows),
                    filled_ordername=filled_ordername,
                    filled_order=filled_order,
                    missing_count=len(missing),
                    skipped_rows=parsed.skipped_rows,
                    missing_orders=missing,
                    duration_ms=int((time.monotonic() - started) * 1000),
                )
                _insert_log(
                    cur,
                    file_name=file_name,
                    file_hash=file_hash,
                    file_size=file_size,
                    result=result,
                )
        return result
    except Exception as exc:
        result = FileProcessResult(
            file_name=file_name,
            status="failed",
            error_message=str(exc),
            duration_ms=int((time.monotonic() - started) * 1000),
        )
        try:
            with local_connection() as conn:
                _ensure_tables(conn)
                with conn.cursor() as cur:
                    _insert_log(
                        cur,
                        file_name=file_name,
                        file_hash=file_hash,
                        file_size=file_size,
                        result=result,
                    )
        except Exception:
            logger.exception("failed to write ozn_upload_log for %s", file_name)
        return result
    finally:
        try:
            path.unlink(missing_ok=True)
        except OSError as exc:
            logger.warning("could not delete %s: %s", path, exc)


def run() -> None:
    inbox = EXCEL_UPLOAD_DIR
    inbox.mkdir(parents=True, exist_ok=True)
    files = list_ready_files(inbox)
    if not files:
        logger.debug("%s: no ready files in %s", JOB_NAME, inbox)
        return

    run_id = None
    with local_connection() as conn:
        _ensure_tables(conn)
        run_id = log_job_run(
            conn,
            JOB_NAME,
            "running",
            f"Found {len(files)} Excel file(s) in {inbox}",
        )

    processed = 0
    failed = 0
    filled = 0
    try:
        for path in files:
            logger.info("%s: processing %s", JOB_NAME, path.name)
            result = process_file(path)
            processed += 1
            filled += result.filled_ordername + result.filled_order
            if result.status == "failed":
                failed += 1
                logger.warning(
                    "%s: %s failed: %s",
                    JOB_NAME,
                    path.name,
                    result.error_message,
                )
            else:
                logger.info(
                    "%s: %s rows=%s ordername=%s order=%s missing=%s skipped=%s",
                    JOB_NAME,
                    path.name,
                    result.excel_rows,
                    result.filled_ordername,
                    result.filled_order,
                    result.missing_count,
                    result.skipped_rows,
                )
    except Exception as exc:
        logger.exception("%s job failed", JOB_NAME)
        with local_connection() as conn:
            log_job_run(
                conn,
                JOB_NAME,
                "failed",
                str(exc),
                rows_affected=filled,
                run_id=run_id,
            )
        raise

    status = "failed" if failed and not (processed - failed) else "success"
    message = (
        f"{processed} file(s), filled={filled}, failed_files={failed}"
    )
    with local_connection() as conn:
        log_job_run(
            conn,
            JOB_NAME,
            status,
            message,
            rows_affected=filled,
            run_id=run_id,
        )
    logger.info("%s finished: %s", JOB_NAME, message)
